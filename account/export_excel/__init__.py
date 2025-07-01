import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from ..models import *


def generate_excel_file(request, order_id):
    try:
        order = Order.objects.get(id=order_id)
    except Order.DoesNotExist:
        return None

    def get_display_value(value):
        if isinstance(value, bool):
            return 'да' if value else 'нет'
        if value is None:
            return 'нет'
        if isinstance(value, str) and not value.strip():
            return 'нет'
        return value

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Order Data'

    header_cell = ws.cell(row=1, column=1, value='Информация по заказу')
    header_cell.font = Font(bold=True)
    header_cell.border = Border(right=Side(style='thick'), bottom=Side(style='thick'))
    header_cell.alignment = Alignment(wrap_text=True)

    order_data = [
        ('Номер заказа', order.order_number),
        ('Наименование предприятия заказчика', order.creator.company_name if hasattr(order.creator, 'company_name') else ''),
        ('Код заказчика', str(order.creator.id)),
        ('Номер шаблона', order.mask_name if order.mask_name else 'нет'),
        ('Техпроцесс', order.technical_process.name_process),
        ('Производственная площадка', order.platform_code.platform_name),
        ('Тип запуска', order.get_order_type_display()),
        ('Число проектов в кадре', str(order.product_count)),
        ('Тип толщины подложки', order.get_substrate_type_display()),
        ('Толщина подложки', f'{order.selected_thickness.value}'),
        ('Диаметр подложки', f'{order.selected_diameter.value} мм'),
        ('Экспериментальная структура', get_display_value(order.experimental_structure)),
        ('Контроль электрических параметров на пластине', order.get_dc_rf_probing_e_map_display()),
        ('Маркировка бракованных по электрическим параметрам', get_display_value(order.dc_rf_probing_inking)),
        ('Визуальный контроль и маркировка брака', get_display_value(order.visual_inspection_inking)),
        ('Предоставление данных контроля параметрического монитора', get_display_value(order.parametric_monitor_control)),
        ('Способ разделения пластины на кристаллы', order.get_dicing_method_display()),
        ('УФ засветка полимерного носителя', get_display_value(order.tape_uv_support)),
        ('Вид поставки пластин', order.get_wafer_deliver_format_display()),
        ('Вид тары для кристаллов', order.get_container_for_crystals_display()),
        ('Схема разделения пластины на кристаллы по схеме заказчика', get_display_value(order.multiplan_dicing_plan)),
        ('Корпусирование силами производителя', get_display_value(order.package_servce)),
        ('Ускоренный запуск производства фотошаблонов', get_display_value(order.delivery_premium_template)),
        ('Ускоренный запуск производства пластин', get_display_value(order.delivery_premium_plate)),
        ('Примечания', order.special_note if order.special_note else 'нет'),
        ('Форму заполнил', request.user.email),
    ]

    for row_num, (header, value) in enumerate(order_data, start=3):
        ws.cell(row=row_num, column=1, value=header).alignment = Alignment(wrap_text=True, horizontal='left')
        ws.cell(row=row_num, column=2, value=value).alignment = Alignment(wrap_text=True, horizontal='left')

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35

    return wb
